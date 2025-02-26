# excel_processor.py
import os
import datetime
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from utils import transform_measurement_method

def select_folder_and_excel():
    """
    利用 tkinter 選取包含 Excel 檔案的資料夾，
    並確認該資料夾內僅有一個 Excel 檔案。
    """
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    folder_path = filedialog.askdirectory(title="選擇包含Excel檔案的資料夾")
    if not folder_path:
        print("未選擇資料夾，程式結束。")
        exit()
    excel_files = [f for f in os.listdir(folder_path) if f.lower().endswith((".xlsx", ".xls"))]
    if len(excel_files) != 1:
        print("資料夾內必須且僅有一個 Excel 檔案，程式結束。")
        exit()
    excel_file_path = os.path.join(folder_path, excel_files[0])
    print("選取的 Excel 檔案：", excel_file_path)
    return folder_path, excel_file_path

def process_excel_pandas(excel_file_path):
    """
    利用 pandas 讀取 Excel 首個工作表前 2 行資料，
    並進行欄位重新命名與資料前置處理。
    """
    xls = pd.ExcelFile(excel_file_path)
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], usecols="A:AA", nrows=2)
    column_mapping = {
        "案號": "case_number",
        "施測日期": "measurement_date",
        "施測人員姓名": "surveyors_name",
        "施測方式": "measurement_method",
        "施測廠商名稱": "survey_company_name",
        "施測廠商電話": "survey_company_phone",
        "技師證號": "technician_license_number",
        "技術士證號": "technician_certificate_number",
        "施測儀器": "survey_equipment",
        "GPS 廠牌型號": "gps_brand_model",
        "經緯儀/全站儀廠牌型號": "total_station_brand_model",
        "潛盾施工廠牌型號": "shield_machine_brand_model",
        "其它廠牌型號": "other_equipment_brand_model",
        "施測點數": "survey_point_count",
        "管線點位": "pipeline_point_count",
        "孔蓋點位": "manhole_point_count",
        "設施物點位": "facility_point_count",
        "參考點位編號": "reference_point_number",
        "參考點位來源": "reference_point_source",
        "原始 E 座標": "original_easting",
        "原始 N 座標": "original_northing",
        "原始 H 正高": "original_height",
        "檢測 E 座標": "measured_easting",
        "檢測 N 座標": "measured_northing",
        "檢測 H 正高": "measured_height",
        "監工名稱": "supervisor_name",
        "區處": "district",
    }
    df_renamed = df.rename(columns=column_mapping)
    df_renamed["measurement_date"] = pd.to_datetime(df_renamed["measurement_date"], errors="coerce")
    df_renamed["measurement_date"] = df_renamed["measurement_date"].apply(
        lambda x: {"year": x.year, "month": x.month, "day": x.day} if pd.notnull(x) else None
    )
    current_date = datetime.datetime.now()
    df_renamed["current_year"]= current_date.year
    df_renamed["current_month"]= current_date.month
    df_renamed["current_day"]= current_date.day

    df_renamed["measurement_method"] = df_renamed["measurement_method"].apply(transform_measurement_method)
    df_renamed["survey_equipment"] = df_renamed["survey_equipment"].apply(transform_measurement_method)
    df_renamed = df_renamed.fillna("empty")
    return df_renamed
def process_excel_openpyxl(excel_file_path):
    """
    使用 openpyxl 讀取 Excel 指定範圍資料，
    並根據 B 欄格式分離為 simulated_data 與 reserved_data。
    此版本改以從第五行起檢查 A~G 欄有資料的列數作為 n_value。
    """
    import openpyxl, re
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active

    start_row = 5
    # 取得從第五行開始、A 至 G 欄的所有列
    all_rows = list(ws.iter_rows(min_row=start_row, max_col=7))
    # 過濾掉完全沒有資料的列
    non_empty_rows = [row for row in all_rows if any(cell.value is not None for cell in row)]
    n_value = len(non_empty_rows)
    
    if n_value == 0:
        print("從第 5 行起在 A 至 G 列沒有發現任何資料。")
        return [], []

    end_row = start_row + n_value - 1
    data_range = ws[f"A{start_row}:G{end_row}"]
    
    pattern = re.compile(r"^\s*\d+管道點\d+-實測\s*$")
    simulated_data = []
    reserved_data = []
    
    for row in data_range:
        b_value = row[1].value
        b_str = str(b_value) if b_value is not None else ""
        if not pattern.match(b_str):
            reserved_data.append({
                "Number": row[0].value,
                "Type": row[1].value,
                "Coordinate_X": round(row[2].value, 3),
                "Coordinate_Y": round(row[3].value, 3),
                "Ground_Elevation": round(row[4].value, 3),
                "Pipe_Burial_Depth": round(row[5].value, 2),
                "Pipe_Top_Coordinate_Z": round(row[4].value, 3),
            })
            continue
        simulated_data.append({
            "Number": row[0].value,
            "Type": row[1].value,
            "Coordinate_X": round(row[2].value, 3),
            "Coordinate_Y": round(row[3].value, 3),
            "Ground_Elevation": round(row[4].value, 3),
            "Pipe_Burial_Depth": round(row[5].value, 2),
            "Pipe_Top_Coordinate_Z": round(row[6].value, 3),
        })
    
    if reserved_data:
        print("以下資料不符合格式，將保留起來，不加入主要表格：")
        for item in reserved_data:
            print(item)
    
    return simulated_data, reserved_data

def create_output_folder(case_number):
    """根據案號建立輸出資料夾，並回傳路徑"""
    output_folder = os.path.join("output", str(case_number))
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    return output_folder
